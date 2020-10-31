from typing import List, Dict, Tuple
from tkinter import messagebox, filedialog
import openpyxl
from sys import exit
from warnings import warn

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


class SalientFeaturePosition:
    featureNumber: int
    positions: Dict[int, Tuple[int, int]]

    def __init__(self, featurenumber: int):
        self.featureNumber = featurenumber
        self.framecount = 0  #
        self.positions = {}  # {frame, (X, Y)}

    def add(self, frame, x, y):
        self.positions[frame] = (x, y)
        self.framecount += 1

    def __gt__(self, other):
        return self.featureNumber > other.featureNumber

    def __lt__(self, other):
        return self.featureNumber < other.featureNumber

    def __eq__(self, other):
        return self.featureNumber == other.featureNumber


class SalientFeatureParser:
    featureCount: int
    sheet: Worksheet
    features: List[SalientFeaturePosition]
    frameList: List[int]

    class ParseFailException(Exception):
        def __init__(self, why):
            self.why = why

    def __init__(self, filepath=None):
        self.featureCount = 0
        self.features = []
        self.frameList = []

        try:
            self.loadsheet(filepath)
            self.countfeatures()
            self.initializefeatures()
            self.importfeatures()
            # Sort for future use
            # It should already be sorted in order of ascending frames
            # But this ensures it
            self.frameList.sort()
        except self.ParseFailException as exp:
            print("File Parsing Failure!")
            print(f"Reason: {exp.why}")
            exit(-1)

    def loadsheet(self, filepath=None):
        if filepath is None:
            file = filedialog.askopenfile(mode='rb', title='Choose a file')
            if file is None:
                messagebox.showerror("File error!", "Error: No file selected.")
                exit(-1)
            filepath = file.name
            # We, uh, don't actually need this file handle. Just its path.
            file.close()
        spreadsheet = openpyxl.load_workbook(filepath, data_only=True)
        sheet = spreadsheet.active
        self.sheet = sheet

    def countfeatures(self):
        for i, column in enumerate(self.sheet.iter_cols()):
            if i == 0:
                continue  # Frame position
            if i % 2 == 0:
                self.featureCount += 1  # Increment for every x, y

    def initializefeatures(self):
        for i in range(self.featureCount):
            self.features.append(SalientFeaturePosition(i))

    def importfeature(self, featurenumber: int, framecell: Cell, xcell: Cell, ycell: Cell):
        frame = framecell.value
        x = xcell.value
        y = ycell.value

        if type(frame) not in (float, int):
            raise self.ParseFailException(f"Frame not a number on row {framecell.row}")
        if type(x) not in (float, int) and x is not None:
            raise self.ParseFailException(f"X position not a number on row {xcell.row}, column {xcell.column}")
        # More than anything, this is a jab at our research team, who in many cases provided extremely incomplete
        # salient feature data, with sometimes hundreds of positions missing or
        # even only one or two positions provided at all.
        if x is None:
            warn(f"Missing X position on row {xcell.row}, column {xcell.column}")
        if type(y) not in (float, int) and y is not None:
            raise self.ParseFailException(f"Y position not a number on row {y.row}, column {y.column}")
        if y is None:
            warn(f"Missing X position on row {ycell.row}, column {ycell.column}", category=RuntimeWarning)

        self.features[featurenumber].add(int(frame), x, y)

    def importfeatures(self):
        for i, featureNumber in zip(range(0, self.featureCount * 2, 2), range(self.featureCount)):
            # frameCol = self.sheet.columns[0]
            # xCol = self.sheet.columns[1 + i]
            # yCol = self.sheet.columns[1 + i + 1]  # Skip frame position, then skip X column
            for row in self.sheet.iter_rows(min_row=2):  # Row 2 for skipping titles
                frame = row[0]
                if frame.value not in self.frameList:
                    self.frameList.append(int(frame.value))
                x = row[1 + i]
                y = row[1 + i + 1]  # Skip frame position, then skip X column
                self.importfeature(featureNumber, frame, x, y)


if __name__ == "__main__":
    parser = SalientFeatureParser()
    pass
